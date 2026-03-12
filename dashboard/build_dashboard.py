from __future__ import annotations

import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

INPUT_XLSX = Path('/Users/anderson/Downloads/NJ AI Policies Schools List.xlsx')
OUTPUT_HTML = Path('/Users/anderson/Desktop/AI Policies folder/dashboard/index.html')

FIELDS = [
    'ID',
    'District Name',
    'County',
    'Minority Enrollment',
    'Below Poverty',
    'SNAP',
    'District Website',
    'Relevant AI Policy Documents',
    'Outside Search Perameters Documents',
    'Risk/Surveillance Index',
    'Opportunity/Innovation Index',
    'Restriction Index',
    'Guidance Index',
    'Primary AI Policy Frame',
    'Classification Confidence',
    'Index Evidence',
    'NCES District Population (2024-2025)',
    'District Size Class (small<2500, medium 2500-9999, large>=10000)',
]


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(',', '')
    text = re.sub(r'%+$', '', text)
    try:
        return float(text)
    except ValueError:
        return None


def looks_like_url(value):
    if value is None:
        return False
    text = str(value).strip().lower()
    return text.startswith('http://') or text.startswith('https://')


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def load_records(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_index = {str(h).strip(): i + 1 for i, h in enumerate(header_cells) if h is not None}

    records = []
    for row in range(2, ws.max_row + 1):
        district_name = ws.cell(row, header_index['District Name']).value
        if district_name in (None, ''):
            continue

        raw = {}
        for field in FIELDS:
            col = header_index.get(field)
            raw[field] = ws.cell(row, col).value if col else None

        minority = to_number(raw['Minority Enrollment'])
        poverty = to_number(raw['Below Poverty'])
        snap = to_number(raw['SNAP'])
        population = to_number(raw['NCES District Population (2024-2025)'])

        relevant_doc = clean_text(raw['Relevant AI Policy Documents'])
        has_policy_doc = bool(relevant_doc and relevant_doc.lower() != 'n/a')

        rec = {
            'id': clean_text(raw['ID']),
            'district': clean_text(raw['District Name']),
            'county': clean_text(raw['County']) or 'Unknown',
            'minorityEnrollment': minority,
            'belowPoverty': poverty,
            'snap': snap,
            'districtWebsite': clean_text(raw['District Website']),
            'policyDoc': relevant_doc,
            'outsideDoc': clean_text(raw['Outside Search Perameters Documents']),
            'riskIndex': clean_text(raw['Risk/Surveillance Index']),
            'innovationIndex': clean_text(raw['Opportunity/Innovation Index']),
            'restrictionIndex': clean_text(raw['Restriction Index']),
            'guidanceIndex': clean_text(raw['Guidance Index']),
            'policyFrame': clean_text(raw['Primary AI Policy Frame']) or 'Unclassified',
            'confidence': clean_text(raw['Classification Confidence']) or 'none',
            'indexEvidence': clean_text(raw['Index Evidence']),
            'population': population,
            'sizeClass': clean_text(raw['District Size Class (small<2500, medium 2500-9999, large>=10000)']) or 'unknown',
            'hasPolicyDoc': has_policy_doc,
            'hasDistrictWebsite': looks_like_url(raw['District Website']),
            'hasClassification': clean_text(raw['Primary AI Policy Frame']) is not None,
        }

        if rec['belowPoverty'] is not None and rec['belowPoverty'] > 1:
            rec['belowPoverty'] = rec['belowPoverty'] / 100.0

        records.append(rec)
    return records


def top_counties(records, limit=15):
    by_county = defaultdict(lambda: {'total': 0, 'classified': 0, 'policy_docs': 0})
    for r in records:
        c = by_county[r['county']]
        c['total'] += 1
        if r['hasClassification']:
            c['classified'] += 1
        if r['hasPolicyDoc']:
            c['policy_docs'] += 1

    ranked = sorted(by_county.items(), key=lambda kv: kv[1]['total'], reverse=True)[:limit]
    out = []
    for county, stats in ranked:
        coverage = (stats['classified'] / stats['total'] * 100) if stats['total'] else 0
        out.append({
            'county': county,
            'totalDistricts': stats['total'],
            'classifiedDistricts': stats['classified'],
            'policyDocDistricts': stats['policy_docs'],
            'classificationCoverage': round(coverage, 1),
        })
    return out


def policy_frame_distribution(records):
    frames = Counter(r['policyFrame'] for r in records if r['hasClassification'])
    ordered = frames.most_common()
    return [{'frame': name, 'count': count} for name, count in ordered]


def summary(records):
    total = len(records)
    classified = sum(1 for r in records if r['hasClassification'])
    docs = sum(1 for r in records if r['hasPolicyDoc'])
    high_conf = sum(1 for r in records if str(r['confidence']).lower() == 'high')
    pop_known = [r['population'] for r in records if r['population'] is not None]
    mean_pop = round(sum(pop_known) / len(pop_known), 1) if pop_known else None
    return {
        'totalDistricts': total,
        'classifiedDistricts': classified,
        'classificationCoverage': round((classified / total) * 100, 1) if total else 0,
        'districtsWithPolicyDocs': docs,
        'policyDocCoverage': round((docs / total) * 100, 1) if total else 0,
        'highConfidenceClassifications': high_conf,
        'averagePopulation': mean_pop,
    }


def build_html(payload):
    data_json = json.dumps(payload, ensure_ascii=True)
    template = """<!doctype html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
  <title>NJ District AI Policy Dashboard</title>
  <style>
    :root {{
      --bg: #f3f6f4;
      --surface: #ffffff;
      --ink: #132125;
      --muted: #5c6f73;
      --accent: #0a7a86;
      --accent-2: #db7c26;
      --grid: #dce4e1;
      --ok: #287f3b;
      --warn: #9b5a00;
      --radius: 14px;
      --shadow: 0 10px 28px rgba(16, 30, 34, 0.08);
    }}

    * {{ box-sizing: border-box; }}

    body {{
      margin: 0;
      font-family: "Avenir Next", "Segoe UI", "Helvetica Neue", sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at 10% 10%, #e6f3ef 0, transparent 30%),
        radial-gradient(circle at 85% 20%, #fcefdc 0, transparent 35%),
        var(--bg);
    }}

    .layout {{
      display: grid;
      grid-template-columns: 280px 1fr;
      min-height: 100vh;
      gap: 18px;
      padding: 18px;
    }}

    .panel {{
      background: var(--surface);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      border: 1px solid #e8efed;
    }}

    .filters {{ padding: 16px; position: sticky; top: 18px; height: fit-content; }}
    .content {{ padding: 0; overflow: hidden; }}

    h1 {{ margin: 0; font-size: 1.45rem; letter-spacing: 0.2px; }}
    h2 {{ margin: 0 0 8px 0; font-size: 1rem; }}
    .sub {{ color: var(--muted); font-size: 0.9rem; margin-top: 6px; }}

    .top {{
      padding: 18px;
      border-bottom: 1px solid var(--grid);
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
      flex-wrap: wrap;
    }}

    .pill {{
      border-radius: 999px;
      padding: 6px 12px;
      font-size: 0.82rem;
      color: #0d5a62;
      background: #def2f5;
      border: 1px solid #bfe3e9;
    }}

    .kpis {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(175px, 1fr));
      gap: 10px;
      padding: 14px 18px 0;
    }}

    .kpi {{
      background: #fafdfc;
      border: 1px solid #e0ebe8;
      border-radius: 12px;
      padding: 12px;
      min-height: 82px;
    }}

    .kpi .label {{ color: var(--muted); font-size: 0.8rem; text-transform: uppercase; letter-spacing: 0.3px; }}
    .kpi .value {{ font-size: 1.55rem; font-weight: 700; margin-top: 6px; }}

    .grid {{
      display: grid;
      grid-template-columns: repeat(12, 1fr);
      gap: 12px;
      padding: 14px 18px 18px;
    }}

    .card {{
      background: #fff;
      border: 1px solid #e3ece9;
      border-radius: 12px;
      padding: 12px;
    }}

    .span-7 {{ grid-column: span 7; }}
    .span-5 {{ grid-column: span 5; }}
    .span-12 {{ grid-column: span 12; }}

    .control {{ margin: 12px 0; }}
    .control label {{ display: block; font-size: 0.8rem; color: var(--muted); margin-bottom: 6px; }}

    input[type=\"text\"], select {{
      width: 100%;
      border: 1px solid #cfddda;
      border-radius: 10px;
      padding: 10px 12px;
      font-size: 0.93rem;
      background: #fff;
      color: var(--ink);
    }}

    button {{
      width: 100%;
      border: 1px solid #bfd6d2;
      border-radius: 10px;
      padding: 9px 12px;
      background: #eef7f6;
      color: #0e4f55;
      cursor: pointer;
      font-weight: 600;
    }}

    button:hover {{ background: #e4f2f1; }}

    .legend {{ display: flex; gap: 12px; flex-wrap: wrap; font-size: 0.8rem; color: var(--muted); margin-bottom: 8px; }}
    .swatch {{ display: inline-block; width: 11px; height: 11px; border-radius: 3px; margin-right: 6px; vertical-align: middle; }}

    .chart {{ width: 100%; height: 320px; border: 1px dashed #d8e4e1; border-radius: 10px; overflow: hidden; }}

    .table-wrap {{ overflow: auto; border: 1px solid #e1ebe8; border-radius: 10px; }}

    table {{ width: 100%; border-collapse: collapse; font-size: 0.9rem; }}
    thead th {{ text-align: left; background: #f4f9f7; color: #41585d; padding: 9px; border-bottom: 1px solid #dfebe7; position: sticky; top: 0; }}
    tbody td {{ padding: 8px 9px; border-bottom: 1px solid #eff5f3; vertical-align: top; }}
    tbody tr:hover {{ background: #fbfefd; }}

    .badge {{
      display: inline-block;
      border-radius: 999px;
      padding: 4px 8px;
      border: 1px solid #d8e6e3;
      background: #f5fbfa;
      font-size: 0.76rem;
      margin-right: 4px;
      margin-bottom: 4px;
    }}

    .match {{ color: var(--ok); border-color: #cae5d0; background: #edf8f0; }}
    .no_match {{ color: var(--warn); border-color: #f0dcc1; background: #fdf5ea; }}

    @media (max-width: 1050px) {{
      .layout {{ grid-template-columns: 1fr; }}
      .filters {{ position: static; }}
      .span-7, .span-5 {{ grid-column: span 12; }}
    }}
  </style>
</head>
<body>
  <div class=\"layout\">
    <aside class=\"panel filters\">
      <h2>Filter & Search</h2>
      <div class=\"sub\">Quick district-level lookups and cross-county comparisons.</div>

      <div class=\"control\">
        <label for=\"districtSearch\">District Name</label>
        <input id=\"districtSearch\" type=\"text\" placeholder=\"Type district name...\" />
      </div>

      <div class=\"control\">
        <label for=\"countySelect\">County</label>
        <select id=\"countySelect\"></select>
      </div>

      <div class=\"control\">
        <label for=\"sizeSelect\">District Size</label>
        <select id=\"sizeSelect\">
          <option value=\"all\">All sizes</option>
          <option value=\"small\">Small</option>
          <option value=\"medium\">Medium</option>
          <option value=\"large\">Large</option>
          <option value=\"unknown\">Unknown</option>
        </select>
      </div>

      <div class=\"control\">
        <label for=\"frameSelect\">Policy Frame</label>
        <select id=\"frameSelect\"></select>
      </div>

      <div class=\"control\">
        <label for=\"classificationSelect\">Classification</label>
        <select id=\"classificationSelect\">
          <option value=\"all\">All</option>
          <option value=\"classified\">Classified only</option>
          <option value=\"unclassified\">Unclassified only</option>
        </select>
      </div>

      <div class=\"control\">
        <button id=\"resetBtn\">Reset Filters</button>
      </div>
    </aside>

    <main class=\"panel content\">
      <div class=\"top\">
        <div>
          <h1>NJ AI Policy Dashboard (District Level)</h1>
          <div class=\"sub\">Data source: NJ AI Policies Schools List.xlsx</div>
        </div>
        <div class=\"pill\" id=\"lastUpdated\"></div>
      </div>

      <section class=\"kpis\" id=\"kpiContainer\"></section>

      <section class=\"grid\">
        <article class=\"card span-7\">
          <h2>Top Counties by District Count</h2>
          <div class=\"sub\">Bars show district totals; line marks classification coverage.</div>
          <div class=\"chart\" id=\"countyChart\"></div>
        </article>

        <article class=\"card span-5\">
          <h2>Policy Frame Distribution</h2>
          <div class=\"sub\">Only districts with a classification are included.</div>
          <div class=\"chart\" id=\"frameChart\"></div>
        </article>

        <article class=\"card span-12\">
          <h2>District Detail Table</h2>
          <div class=\"sub\">Use the search box to instantly pull district-level records.</div>
          <div class=\"table-wrap\">
            <table>
              <thead>
                <tr>
                  <th>District</th>
                  <th>County</th>
                  <th>Size</th>
                  <th>Population</th>
                  <th>Policy Frame</th>
                  <th>Confidence</th>
                  <th>Indexes</th>
                  <th>Links</th>
                </tr>
              </thead>
              <tbody id=\"districtTbody\"></tbody>
            </table>
          </div>
        </article>
      </section>
    </main>
  </div>

  <script>
    const DASHBOARD_DATA = __DATA_JSON__;

    const state = {{
      districtSearch: '',
      county: 'all',
      sizeClass: 'all',
      frame: 'all',
      classification: 'all',
    }};

    const el = {{
      districtSearch: document.getElementById('districtSearch'),
      countySelect: document.getElementById('countySelect'),
      sizeSelect: document.getElementById('sizeSelect'),
      frameSelect: document.getElementById('frameSelect'),
      classificationSelect: document.getElementById('classificationSelect'),
      resetBtn: document.getElementById('resetBtn'),
      kpiContainer: document.getElementById('kpiContainer'),
      countyChart: document.getElementById('countyChart'),
      frameChart: document.getElementById('frameChart'),
      districtTbody: document.getElementById('districtTbody'),
      lastUpdated: document.getElementById('lastUpdated'),
    }};

    const fmtInt = new Intl.NumberFormat('en-US', {{ maximumFractionDigits: 0 }});
    const fmtPct = new Intl.NumberFormat('en-US', {{ style: 'percent', maximumFractionDigits: 1 }});

    function unique(arr) {{
      return [...new Set(arr.filter(Boolean))].sort((a, b) => a.localeCompare(b));
    }}

    function buildSelect(select, values, allLabel = 'All') {{
      select.innerHTML = '';
      const all = document.createElement('option');
      all.value = 'all';
      all.textContent = allLabel;
      select.appendChild(all);
      values.forEach(v => {{
        const opt = document.createElement('option');
        opt.value = v;
        opt.textContent = v;
        select.appendChild(opt);
      }});
    }}

    function safeText(v, fallback = 'n/a') {{
      return v === null || v === undefined || v === '' ? fallback : String(v);
    }}

    function compactPct(value) {{
      if (value === null || value === undefined || Number.isNaN(value)) return 'n/a';
      return fmtPct.format(value <= 1 ? value : value / 100);
    }}

    function statusBadge(value) {{
      if (!value) return '<span class="badge">n/a</span>';
      const cls = value === 'match' ? 'match' : value === 'no_match' ? 'no_match' : '';
      return `<span class=\"badge ${cls}\">${value}</span>`;
    }}

    function linksCell(record) {{
      const links = [];
      if (record.districtWebsite && record.districtWebsite.startsWith('http')) {{
        links.push(`<a href=\"${record.districtWebsite}\" target=\"_blank\" rel=\"noopener noreferrer\">District Site</a>`);
      }}
      if (record.policyDoc && record.policyDoc.startsWith('http')) {{
        links.push(`<a href=\"${record.policyDoc}\" target=\"_blank\" rel=\"noopener noreferrer\">Policy Doc</a>`);
      }}
      if (record.outsideDoc && record.outsideDoc.startsWith('http')) {{
        links.push(`<a href=\"${record.outsideDoc}\" target=\"_blank\" rel=\"noopener noreferrer\">Outside Doc</a>`);
      }}
      return links.length ? links.join('<br/>') : 'n/a';
    }}

    function filteredRecords() {{
      const q = state.districtSearch.toLowerCase();
      return DASHBOARD_DATA.records.filter(r => {{
        if (q && !r.district.toLowerCase().includes(q)) return false;
        if (state.county !== 'all' && r.county !== state.county) return false;
        if (state.sizeClass !== 'all' && r.sizeClass !== state.sizeClass) return false;
        if (state.frame !== 'all' && r.policyFrame !== state.frame) return false;
        if (state.classification === 'classified' && !r.hasClassification) return false;
        if (state.classification === 'unclassified' && r.hasClassification) return false;
        return true;
      }});
    }}

    function computeSummary(records) {{
      const total = records.length;
      const classified = records.filter(r => r.hasClassification).length;
      const docs = records.filter(r => r.hasPolicyDoc).length;
      const high = records.filter(r => (r.confidence || '').toLowerCase() === 'high').length;
      const pop = records.filter(r => typeof r.population === 'number').map(r => r.population);
      const avgPop = pop.length ? pop.reduce((a, b) => a + b, 0) / pop.length : null;

      return {{
        total,
        classified,
        classifiedPct: total ? classified / total : 0,
        docs,
        docsPct: total ? docs / total : 0,
        high,
        avgPop,
      }};
    }}

    function renderKPIs(records) {{
      const s = computeSummary(records);
      const cards = [
        ['Districts (filtered)', fmtInt.format(s.total)],
        ['Classified', `${fmtInt.format(s.classified)} (${fmtPct.format(s.classifiedPct)})`],
        ['With Policy Docs', `${fmtInt.format(s.docs)} (${fmtPct.format(s.docsPct)})`],
        ['High Confidence', fmtInt.format(s.high)],
        ['Avg Population', s.avgPop ? fmtInt.format(s.avgPop) : 'n/a'],
      ];

      el.kpiContainer.innerHTML = cards.map(([label, value]) => `
        <div class=\"kpi\">
          <div class=\"label\">${label}</div>
          <div class=\"value\">${value}</div>
        </div>
      `).join('');
    }}

    function renderCountyChart(records) {{
      const byCounty = new Map();
      records.forEach(r => {{
        const curr = byCounty.get(r.county) || {{ total: 0, classified: 0 }};
        curr.total += 1;
        if (r.hasClassification) curr.classified += 1;
        byCounty.set(r.county, curr);
      }});

      const points = [...byCounty.entries()]
        .map(([county, s]) => ({{ county, ...s, coverage: s.total ? s.classified / s.total : 0 }}))
        .sort((a, b) => b.total - a.total)
        .slice(0, 15);

      const width = el.countyChart.clientWidth || 640;
      const height = el.countyChart.clientHeight || 320;
      const pad = {{ t: 20, r: 45, b: 78, l: 45 }};
      const innerW = width - pad.l - pad.r;
      const innerH = height - pad.t - pad.b;

      const maxTotal = Math.max(...points.map(p => p.total), 1);
      const barW = points.length ? innerW / points.length : innerW;

      let svg = `<svg viewBox=\"0 0 ${width} ${height}\" width=\"100%\" height=\"100%\">`;
      svg += `<line x1=\"${pad.l}\" y1=\"${pad.t + innerH}\" x2=\"${pad.l + innerW}\" y2=\"${pad.t + innerH}\" stroke=\"#9db3b0\"/>`;

      for (let i = 0; i <= 4; i++) {{
        const y = pad.t + (innerH * i / 4);
        const val = Math.round(maxTotal * (1 - i / 4));
        svg += `<line x1=\"${pad.l}\" y1=\"${y}\" x2=\"${pad.l + innerW}\" y2=\"${y}\" stroke=\"#e3eeeb\"/>`;
        svg += `<text x=\"${pad.l - 8}\" y=\"${y + 4}\" text-anchor=\"end\" font-size=\"10\" fill=\"#5c6f73\">${val}</text>`;
      }}

      const linePts = [];
      points.forEach((p, idx) => {{
        const h = (p.total / maxTotal) * innerH;
        const x = pad.l + idx * barW + barW * 0.16;
        const y = pad.t + innerH - h;
        const bw = barW * 0.68;

        svg += `<rect x=\"${x}\" y=\"${y}\" width=\"${bw}\" height=\"${h}\" fill=\"#0a7a86\" rx=\"4\"/>`;
        svg += `<text x=\"${x + bw/2}\" y=\"${pad.t + innerH + 12}\" font-size=\"10\" fill=\"#30474c\" text-anchor=\"end\" transform=\"rotate(-35 ${x + bw/2} ${pad.t + innerH + 12})\">${p.county}</text>`;
        svg += `<text x=\"${x + bw/2}\" y=\"${y - 4}\" font-size=\"10\" text-anchor=\"middle\" fill=\"#0f2d31\">${p.total}</text>`;

        const lx = x + bw / 2;
        const ly = pad.t + innerH - (p.coverage * innerH);
        linePts.push([lx, ly, p.coverage]);
      }});

      if (linePts.length) {{
        const path = linePts.map((pt, i) => `${i === 0 ? 'M' : 'L'} ${pt[0]} ${pt[1]}`).join(' ');
        svg += `<path d=\"${path}\" fill=\"none\" stroke=\"#db7c26\" stroke-width=\"2.5\"/>`;
        linePts.forEach(([x, y, cov]) => {{
          svg += `<circle cx=\"${x}\" cy=\"${y}\" r=\"3.2\" fill=\"#db7c26\"/>`;
          svg += `<title>${Math.round(cov * 100)}% classified</title>`;
        }});
      }}

      svg += `<text x=\"${width - 6}\" y=\"12\" text-anchor=\"end\" font-size=\"10\" fill=\"#db7c26\">Classification coverage</text>`;
      svg += '</svg>';

      el.countyChart.innerHTML = `
        <div class=\"legend\">
          <span><span class=\"swatch\" style=\"background:#0a7a86\"></span>District count</span>
          <span><span class=\"swatch\" style=\"background:#db7c26\"></span>Classification coverage</span>
        </div>
        ${svg}
      `;
    }}

    function renderFrameChart(records) {{
      const counts = new Map();
      records.filter(r => r.hasClassification).forEach(r => {{
        counts.set(r.policyFrame, (counts.get(r.policyFrame) || 0) + 1);
      }});

      const frames = [...counts.entries()].sort((a, b) => b[1] - a[1]);
      const total = frames.reduce((sum, [, n]) => sum + n, 0) || 1;

      const width = el.frameChart.clientWidth || 500;
      const height = el.frameChart.clientHeight || 320;
      const pad = {{ t: 18, r: 20, b: 18, l: 20 }};
      const innerW = width - pad.l - pad.r;
      const innerH = height - pad.t - pad.b;
      const rowH = Math.max(26, Math.min(44, innerH / Math.max(frames.length, 1)));

      const palette = ['#0a7a86', '#db7c26', '#3e8b45', '#007f6d', '#8c4b9b', '#1665ad', '#7f6a2a'];

      let svg = `<svg viewBox=\"0 0 ${width} ${height}\" width=\"100%\" height=\"100%\">`;

      frames.forEach(([frame, count], i) => {{
        const y = pad.t + i * rowH + 2;
        const barMaxW = innerW * 0.55;
        const w = Math.max(2, (count / total) * barMaxW);
        const color = palette[i % palette.length];

        svg += `<text x=\"${pad.l}\" y=\"${y + 13}\" font-size=\"11\" fill=\"#31464b\">${frame}</text>`;
        svg += `<rect x=\"${pad.l + innerW * 0.42}\" y=\"${y}\" width=\"${w}\" height=\"14\" fill=\"${color}\" rx=\"4\"/>`;
        svg += `<text x=\"${pad.l + innerW * 0.42 + w + 6}\" y=\"${y + 12}\" font-size=\"11\" fill=\"#1f3337\">${count} (${Math.round(count / total * 100)}%)</text>`;
      }});

      if (!frames.length) {{
        svg += `<text x=\"${width/2}\" y=\"${height/2}\" text-anchor=\"middle\" fill=\"#5c6f73\">No classified districts in current filter.</text>`;
      }}

      svg += '</svg>';
      el.frameChart.innerHTML = svg;
    }}

    function renderTable(records) {{
      const sorted = [...records].sort((a, b) => a.district.localeCompare(b.district)).slice(0, 250);
      el.districtTbody.innerHTML = sorted.map(r => `
        <tr>
          <td><strong>${safeText(r.district)}</strong></td>
          <td>${safeText(r.county)}</td>
          <td>${safeText(r.sizeClass)}</td>
          <td>${r.population !== null ? fmtInt.format(r.population) : 'n/a'}</td>
          <td>${safeText(r.policyFrame, 'Unclassified')}</td>
          <td>${safeText(r.confidence)}</td>
          <td>
            ${statusBadge(r.riskIndex)}
            ${statusBadge(r.innovationIndex)}
            ${statusBadge(r.restrictionIndex)}
            ${statusBadge(r.guidanceIndex)}
          </td>
          <td>${linksCell(r)}</td>
        </tr>
      `).join('') || `<tr><td colspan=\"8\" style=\"text-align:center; color:#5c6f73;\">No districts match current filters.</td></tr>`;
    }}

    function renderAll() {{
      const rows = filteredRecords();
      renderKPIs(rows);
      renderCountyChart(rows);
      renderFrameChart(rows);
      renderTable(rows);
    }}

    function initialize() {{
      const counties = unique(DASHBOARD_DATA.records.map(r => r.county));
      const frames = unique(DASHBOARD_DATA.records.filter(r => r.hasClassification).map(r => r.policyFrame));

      buildSelect(el.countySelect, counties, 'All counties');
      buildSelect(el.frameSelect, frames, 'All policy frames');

      el.lastUpdated.textContent = `Updated ${DASHBOARD_DATA.generatedAt}`;

      el.districtSearch.addEventListener('input', () => {{ state.districtSearch = el.districtSearch.value.trim(); renderAll(); }});
      el.countySelect.addEventListener('change', () => {{ state.county = el.countySelect.value; renderAll(); }});
      el.sizeSelect.addEventListener('change', () => {{ state.sizeClass = el.sizeSelect.value; renderAll(); }});
      el.frameSelect.addEventListener('change', () => {{ state.frame = el.frameSelect.value; renderAll(); }});
      el.classificationSelect.addEventListener('change', () => {{ state.classification = el.classificationSelect.value; renderAll(); }});

      el.resetBtn.addEventListener('click', () => {{
        state.districtSearch = '';
        state.county = 'all';
        state.sizeClass = 'all';
        state.frame = 'all';
        state.classification = 'all';

        el.districtSearch.value = '';
        el.countySelect.value = 'all';
        el.sizeSelect.value = 'all';
        el.frameSelect.value = 'all';
        el.classificationSelect.value = 'all';
        renderAll();
      }});

      window.addEventListener('resize', () => {{ renderCountyChart(filteredRecords()); renderFrameChart(filteredRecords()); }});

      renderAll();
    }}

    initialize();
  </script>
</body>
</html>
"""
    template = template.replace('{{', '{').replace('}}', '}')
    return template.replace('__DATA_JSON__', data_json)


def main():
    records = load_records(INPUT_XLSX)
    payload = {
        'generatedAt': __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'),
        'summary': summary(records),
        'topCounties': top_counties(records),
        'frameDistribution': policy_frame_distribution(records),
        'records': records,
    }

    OUTPUT_HTML.write_text(build_html(payload), encoding='utf-8')
    print(f'Wrote dashboard to: {OUTPUT_HTML}')
    print(f'Records: {len(records)}')


if __name__ == '__main__':
    main()
