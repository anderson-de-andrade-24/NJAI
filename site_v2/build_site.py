from __future__ import annotations

import csv
import json
import os
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SITE_ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = SITE_ROOT / "data"
OUTPUT_JSON = OUTPUT_DIR / "district_dataset.json"
OUTPUT_CSV = OUTPUT_DIR / "district_dataset.csv"
OUTPUT_SUMMARY = OUTPUT_DIR / "build_summary.json"
OUTPUT_GEOJSON = OUTPUT_DIR / "nj-school-districts.geojson"

HEADER_ROW = 3
DATA_START_ROW = 4
SHEET_NAME = "District"

WORKBOOK_CANDIDATES = [
    os.environ.get("NJAI_WORKBOOK"),
    str(SITE_ROOT / "5-25_with_ai_policy (1).xlsx"),
    str(SITE_ROOT.parent / "5-25_with_ai_policy (1).xlsx"),
    str(Path.home() / "Downloads" / "5-25_with_ai_policy (1).xlsx"),
    str(SITE_ROOT / "Copy of NJ AI Policies Schools List (3-30).xlsx"),
    str(SITE_ROOT.parent / "Copy of NJ AI Policies Schools List (3-30).xlsx"),
]

RACE_FIELDS = [
    ("white", "White", "%White"),
    ("black", "Black", "%Black"),
    ("hispanic", "Hispanic", "%Hispanic"),
    ("asian", "Asian", "%Asian"),
    ("native_american", "Native American", "%Native American"),
    ("hawaiian_native", "Hawaiian Native", "%Hawaiian Native"),
    ("two_or_more_races", "Two or More Races", "%Two or More Races"),
]

GRADE_FIELDS = [
    ("pre_k_halfday", "Pre-K Halfday", "%Pre-K Halfday"),
    ("pre_k_fullday", "Pre-K FullDay", "%Pre-K FullDay"),
    ("kindergarten_halfday", "Kindergarten Halfday", "%Kindergarten Halfday"),
    ("kindergarten_fullday", "Kindergarten Fullday", "%Kindergarten Fullday"),
    ("first_grade", "First Grade", "%First Grade"),
    ("second_grade", "Second Grade", "%Second Grade"),
    ("third_grade", "Third Grade", "%Third Grade"),
    ("fourth_grade", "Fourth Grade", "%Fourth Grade"),
    ("fifth_grade", "Fifth Grade", "%Fifth Grade"),
    ("sixth_grade", "Sixth Grade", "%Sixth Grade"),
    ("seventh_grade", "Seventh Grade", "%Seventh Grade"),
    ("eighth_grade", "Eighth Grade", "%Eighth Grade"),
    ("ninth_grade", "Ninth Grade", "%Ninth Grade"),
    ("tenth_grade", "Tenth Grade", "%Tenth Grade"),
    ("eleventh_grade", "Eleventh Grade", "%Eleventh Grade"),
    ("twelfth_grade", "Twelfth Grade", "%Twelfth Grade"),
]

SUPPORT_FIELDS = [
    ("free_lunch", "Free Lunch ", "%Free Lunch"),
    ("reduced_lunch", "Reduced Lunch", "%Reduced Lunch"),
    ("multilingual_learners", "Multilingual Learners", "%Multilingual Learners"),
    ("migrant", "Migrant", "%Migrant"),
    ("military", "Military", "%Military"),
    ("homeless", "Homeless", "%Homeless"),
]


def resolve_workbook() -> Path:
    for candidate in WORKBOOK_CANDIDATES:
        if not candidate:
            continue
        path = Path(candidate).expanduser()
        if path.exists():
            return path
    raise FileNotFoundError(
        "Could not locate the source workbook. Set NJAI_WORKBOOK or place "
        "'5-25_with_ai_policy (1).xlsx' near build_site.py."
    )


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"n/a", "na"}:
        return None
    return text


def to_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_int(value):
    number = to_float(value)
    if number is None:
        return None
    return int(round(number))


def to_fraction(value):
    number = to_float(value)
    if number is None:
        return None
    return number / 100.0


def clean_urbanicity(value):
    text = clean_text(value)
    if not text:
        return None
    return text.replace(" (inferred)", "")


def to_bool(value):
    number = to_float(value)
    if number is None:
        text = clean_text(value)
        if text is None:
            return False
        return text.lower() in {"yes", "true", "1", "y"}
    return bool(int(number))


def to_date_string(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except TypeError:
            return clean_text(value)
    return clean_text(value)


def url_or_none(value):
    text = clean_text(value)
    if text and text.lower().startswith(("http://", "https://")):
        return text
    return None


def unique_in_order(values):
    ordered = OrderedDict()
    for value in values:
        if value is None:
            continue
        key = str(value).strip().lower()
        if not key:
            continue
        ordered[key] = str(value).strip()
    return list(ordered.values())


def build_header_index(headers):
    return {str(header).strip(): index for index, header in enumerate(headers) if header is not None}


def value_for(row, header_index, name):
    index = header_index.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


def district_type_from_row(row, header_index):
    serves_primary = to_bool(value_for(row, header_index, "Serves Primary (PK-8)"))
    serves_secondary = to_bool(value_for(row, header_index, "Serves Secondary (9-12)"))
    serves_unified = to_bool(value_for(row, header_index, "Serves Both / Unified"))

    if serves_unified:
        return "U", "Unified"
    if serves_secondary and not serves_primary:
        return "S", "Secondary"
    if serves_primary and not serves_secondary:
        return "E", "Elementary"
    if serves_primary and serves_secondary:
        return "U", "Unified"
    return "UNK", "Not classified"


def make_breakdown(row, header_index, fields):
    payload = {}
    for slug, count_field, pct_field in fields:
        payload[slug] = {
            "label": count_field.strip(),
            "count": to_int(value_for(row, header_index, count_field)),
            "percent": to_fraction(value_for(row, header_index, pct_field)),
        }
    return payload


def compute_frpl_percent(support):
    free_pct = support["free_lunch"]["percent"]
    reduced_pct = support["reduced_lunch"]["percent"]
    if free_pct is None and reduced_pct is None:
        return None
    return min((free_pct or 0) + (reduced_pct or 0), 1)


def build_records():
    workbook_path = resolve_workbook()
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

    headers = list(next(sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True)))
    header_index = build_header_index(headers)

    records = []
    for row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
        district_name = clean_text(value_for(row, header_index, "District Name"))
        if not district_name:
            continue

        district_type_code, district_type_label = district_type_from_row(row, header_index)
        races = make_breakdown(row, header_index, RACE_FIELDS)
        grades = make_breakdown(row, header_index, GRADE_FIELDS)
        support = make_breakdown(row, header_index, SUPPORT_FIELDS)

        ai_policy = to_bool(value_for(row, header_index, "AI Policy?"))
        policy_link = url_or_none(value_for(row, header_index, "Relevant AI Policy Documents"))
        numeric_policy = to_int(value_for(row, header_index, "Policy Number"))
        policy_number = str(numeric_policy) if numeric_policy is not None else clean_text(
            value_for(row, header_index, "Policy Number")
        )

        policy_dates = unique_in_order(
            [
                to_date_string(value_for(row, header_index, "Adopted")),
                to_date_string(value_for(row, header_index, "Revision")),
                to_date_string(value_for(row, header_index, "Revision #2")),
                to_date_string(value_for(row, header_index, "Revision #3")),
            ]
        )

        total_enrollment = to_int(value_for(row, header_index, "Total Enrollment"))
        white_percent = races["white"]["percent"]
        students_of_color_percent = None if white_percent is None else max(0, min(1, 1 - white_percent))

        record = {
            "county_code": clean_text(value_for(row, header_index, "County Code")),
            "county": clean_text(value_for(row, header_index, "County Name")) or "Unknown",
            "district_code": clean_text(value_for(row, header_index, "District Code")),
            "district_name": district_name,
            "district_type": district_type_code,
            "district_type_label": district_type_label,
            "total_enrollment": total_enrollment,
            "number_of_schools": to_int(value_for(row, header_index, "Number of Schools")),
            "urbanicity": clean_urbanicity(value_for(row, header_index, "Urbanicity")),
            "urbanicity_score": to_float(value_for(row, header_index, "Urbanicity Score")),
            "has_ai_policy": ai_policy or bool(policy_link) or bool(policy_number),
            "ai_policy_flag": ai_policy,
            "has_se_policy": to_bool(value_for(row, header_index, "SE?")),
            "policy_link": policy_link,
            "policy_number": policy_number,
            "policy_dates": policy_dates,
            "policy_adopted": to_date_string(value_for(row, header_index, "Adopted")),
            "policy_revision_dates": [
                date
                for date in [
                    to_date_string(value_for(row, header_index, "Revision")),
                    to_date_string(value_for(row, header_index, "Revision #2")),
                    to_date_string(value_for(row, header_index, "Revision #3")),
                ]
                if date
            ],
            "students_of_color_percent": students_of_color_percent,
            "frpl_percent": compute_frpl_percent(support),
            "multilingual_percent": support["multilingual_learners"]["percent"],
            "race_breakdown": races,
            "grade_breakdown": grades,
            "student_support": support,
            "district_profile": {
                "serves_primary": to_bool(value_for(row, header_index, "Serves Primary (PK-8)")),
                "serves_secondary": to_bool(value_for(row, header_index, "Serves Secondary (9-12)")),
                "serves_unified": to_bool(value_for(row, header_index, "Serves Both / Unified")),
            },
        }
        records.append(record)

    records.sort(key=lambda item: item["district_name"])
    return workbook_path, records


def summary_payload(workbook_path: Path, records):
    with_policy = sum(1 for row in records if row["has_ai_policy"])
    unified = sum(1 for row in records if row["district_type"] == "U")
    secondary = sum(1 for row in records if row["district_type"] == "S")
    elementary = sum(1 for row in records if row["district_type"] == "E")
    counties = len({row["county"] for row in records if row["county"]})
    avg_enrollment = round(
        sum(row["total_enrollment"] or 0 for row in records) / max(len(records), 1)
    )

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source_workbook": workbook_path.name,
        "record_count": len(records),
        "districts_with_ai_policy": with_policy,
        "counties": counties,
        "average_enrollment": avg_enrollment,
        "district_type_counts": {
            "unified": unified,
            "secondary": secondary,
            "elementary": elementary,
        },
        "available_sections": [
            "policy",
            "overview",
            "race_breakdown",
            "grade_breakdown",
            "student_support",
            "district_profile",
        ],
    }


def write_outputs(records, summary):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(
        json.dumps({"summary": summary, "records": records}, indent=2),
        encoding="utf-8",
    )

    fieldnames = [
        "county_code",
        "county",
        "district_code",
        "district_name",
        "district_type",
        "district_type_label",
        "total_enrollment",
        "number_of_schools",
        "urbanicity",
        "urbanicity_score",
        "has_ai_policy",
        "ai_policy_flag",
        "has_se_policy",
        "policy_link",
        "policy_number",
        "policy_adopted",
        "policy_revision_dates",
        "students_of_color_percent",
        "frpl_percent",
        "multilingual_percent",
    ]
    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in records:
            flat = {key: row.get(key) for key in fieldnames}
            flat["policy_revision_dates"] = ", ".join(row["policy_revision_dates"])
            writer.writerow(flat)

    if OUTPUT_GEOJSON.exists():
        geojson_text = OUTPUT_GEOJSON.read_text(encoding="utf-8")
        OUTPUT_GEOJSON.write_text(geojson_text, encoding="utf-8")
    OUTPUT_SUMMARY.write_text(json.dumps(summary, indent=2), encoding="utf-8")


def main():
    workbook_path, records = build_records()
    summary = summary_payload(workbook_path, records)
    write_outputs(records, summary)
    print(f"Wrote {len(records)} records to {OUTPUT_JSON}")
    print(f"Source workbook: {workbook_path}")


if __name__ == "__main__":
    main()
